print("Starting import")
from pandas import DataFrame, ExcelWriter, read_csv
from argparse import ArgumentParser
from datetime import datetime
from os import listdir, path as os_path
from copy import deepcopy as copy
from openpyxl import load_workbook

from Tester import tester_constants, load_testing_diagnosers_data

parser = ArgumentParser(description="Run all tests")
parser.add_argument("-o", "--output", type=str, help="Output file name prefix, default is the result_TIMESTAMP", default=f"{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}")
args = parser.parse_args()

group_by_columns = ["dataset name", "after size", "drift size", "drifted features types", "total drift type", "drift severity level"]
common_aggregated_columns = ["after accuracy decrease", "after retrain accuracy increase", "before after retrain accuracy increase"]
diagnoser_aggregated_columns_suffixes = ["fix accuracy increase", "wasted effort", "correctly_identified"]

columns_dtypes = {
    "dataset name": "string",
    "after size": "float64",
    "drift size": "int64",
    "drift severity level": "int64",
    "drift description": "string",
    "drifted features types": "string",
    "total drift type": "string",
    "tree size": "int64",
    "after accuracy decrease": "float64",
    "after retrain accuracy increase": "float64",
    "before after retrain accuracy increase": "float64"
}

diagnoser_dtypes_suffixes = {
    "diagnoses": "string",
    "faulty features": "string",
    "wasted effort": "int64",
    "correctly_identified": "float64",
    "fix accuracy": "float64",
    "fix accuracy increase": "float64"
}

aggregated_columns = copy(common_aggregated_columns)
aggregating_functions_dict = {common_aggregated_column: "sum" for common_aggregated_column in common_aggregated_columns}

diagnosers_data = load_testing_diagnosers_data()
diagnosers_output_names = list(map(lambda diagnoser_data: diagnoser_data["output_name"], diagnosers_data))

for diagnoser_name in diagnosers_output_names:
    for diagnoser_aggregated_column_suffix in diagnoser_aggregated_columns_suffixes:
        diagnoser_aggregated_column = f"{diagnoser_name} {diagnoser_aggregated_column_suffix}"
        aggregated_columns.append(diagnoser_aggregated_column)
        aggregating_functions_dict[diagnoser_aggregated_column] = "sum"
    for diagnoser_column_suffix, diagnoser_column_dtype in diagnoser_dtypes_suffixes.items():
        diagnoser_column = f"{diagnoser_name} {diagnoser_column_suffix}"
        columns_dtypes[diagnoser_column] = diagnoser_column_dtype
aggregated_count_column_name = "drift description"
aggregating_functions_dict[aggregated_count_column_name] = "count"


output_df = DataFrame(columns=group_by_columns + aggregated_columns + ["count"]).set_index(group_by_columns)

for current_file_index, current_file_name in enumerate(listdir(tester_constants.TEMP_RESULTS_FULL_PATH), 1):
    print("Working on file", current_file_index, ":", current_file_name)
    if not current_file_name.startswith(tester_constants.RESULTS_FILE_NAME_PREFIX):
        continue
    
    current_results_df = None
    with open(os_path.join(tester_constants.TEMP_RESULTS_FULL_PATH, current_file_name), "r") as current_file:
        current_results_df = read_csv(current_file, dtype=columns_dtypes)
    current_group_by_df = current_results_df.groupby(group_by_columns).agg(aggregating_functions_dict)
    current_group_by_df.rename(columns={aggregated_count_column_name: "count"}, inplace=True)
    assert all(current_group_by_df.columns == output_df.columns), f"Columns mismatch in {current_file_name}"
    assert current_group_by_df.index.names == output_df.index.names, f"Index names mismatch in {current_file_name}"
    output_df = output_df.add(current_group_by_df, fill_value=0)

output_df = output_df[["count"] + aggregated_columns]

output_full_path = os_path.join(tester_constants.RESULTS_FULL_PATH, f"{tester_constants.RESULTS_FILE_NAME_PREFIX}_{args.output}.xlsx")
merged_results_sheet_name = "merged_results"
excel_writer_arguments = {
    "path": output_full_path,
    "mode": "w",
    "engine": "openpyxl"
}
if os_path.exists(output_full_path):
    excel_writer_arguments["mode"] = "a"
    excel_writer_arguments["if_sheet_exists"] = "replace"
with ExcelWriter(**excel_writer_arguments) as excel_writer:
    output_df.to_excel(excel_writer, sheet_name=merged_results_sheet_name, merge_cells=False)
output_workbook = load_workbook(output_full_path)
output_workbook[merged_results_sheet_name].sheet_view.rightToLeft = True
output_workbook.save(output_full_path)

print(f"Results saved to {output_full_path}")